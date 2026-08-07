import streamlit as st
import pandas as pd
from datetime import datetime, time
import io
import re
import tempfile
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="EcoAudit Pro - Gestão de Resíduos",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- LISTA DE ABAS E NICHOS ---
LISTA_ABAS = [
    "📋 1. Local", 
    "♻️ 2. Mapeamento", 
    "🔎 3. Auditoria", 
    "📸 4. Fotos & Riscos", 
    "📊 5. Score & Multas",
    "📄 6. Documentos",
    "🚀 7. Plano de Ação & Minuta PGRS",
    "💰 8. Calculadora de Orçamento",
    "📋 9. Passo a Passo do Gerenciamento",
    "📜 10. Legislação & Sistemas"
]

NICHOS_DISPONIVEIS = [
    "Saúde / Veterinária", 
    "Oficina / Posto / Funilaria", 
    "Alimentação (Restaurante/Mercado/Hotel)", 
    "Indústria PME", 
    "Construção Civil", 
    "Condomínio / Comercial"
]

# --- FUNÇÃO PARA INICIALIZAR OU RESETAR DADOS ---
def resetar_dados():
    st.session_state.aba_ativa = LISTA_ABAS[0]
    st.session_state.db = {
        "consultor": "Renan",
        "data_vistoria": datetime.now(),
        "empresa": "",
        "cnpj": "",
        "segmento": NICHOS_DISPONIVEIS[5],
        "resp": "",
        "tel": "",
        "email": "",
        "func": 15,
        "area": 350.0,
        "h_ini": time(8, 0),
        "h_fim": time(18, 0),
        "filiais": "Não",
        "prod": "Não",
        "obs": "Operação regular em horário comercial.",
        "residuos_selecionados": [],
        "outros_residuos": "",
        "sabe_qtd": "Não",
        "balanca": "Não",
        "kg_dia": 15.0,
        "kg_mes": 450.0,
        "aud_resp": {},
        "notas": [5.0] * 8,
        "rk_legal": "Alto",
        "rk_amb": "Médio",
        "rk_op": "Médio",
        "nc": "Ausência de PGRS aprovado; Armazenamento temporário de resíduos em local descoberto; Mistura de resíduos recicláveis com orgânicos.",
        "rec": "Elaborar PGRS/PGRSS com ART; Instalar cobertura no abrigo de resíduos; Treinar equipe operacional.",
        "servicos": [],
        "docs": {},
        "foto_memoria": None,
        "v_hora": 120.0,
        "c_extra": 0.0,
        "margem": 30
    }

# =====================================================================
# 🔐 COFRE DE DADOS (IMPEDE O RESET AO MUDAR DE PÁGINA)
# =====================================================================
if "aba_ativa" not in st.session_state or "db" not in st.session_state:
    resetar_dados()

db = st.session_state.db

def formatar_nome_arquivo(empresa_nome, nicho_nome):
    emp_clean = re.sub(r'[^\w\s-]', '', empresa_nome).strip().replace(' ', '_')
    nicho_clean = re.sub(r'[^\w\s-]', '', nicho_nome).strip().replace(' ', '_')
    if not emp_clean: emp_clean = "Empresa"
    return f"{emp_clean}_{nicho_clean}"

def pular_aba(nome_aba):
    st.session_state.aba_ativa = nome_aba

def sincronizar_menu_dropdown():
    st.session_state.aba_ativa = st.session_state.w_select_navegacao

# --- FUNÇÃO DE RECOMENDAÇÃO AUTOMÁTICA DE SERVIÇOS ---
def gerar_servicos_recomendados():
    servicos_sugeridos = set()
    
    seg = db["segmento"]
    if seg == "Saúde / Veterinária":
        servicos_sugeridos.add("Elaboração do PGRSS (Saúde)")
        servicos_sugeridos.add("Treinamento de Biossegurança e Perfurocortantes")
    elif seg == "Construção Civil":
        servicos_sugeridos.add("Elaboração do PGRCC (Obras/Construção)")
        servicos_sugeridos.add("Homologação de Destinadores e ATT/Aterros")
    elif seg == "Oficina / Posto / Funilaria":
        servicos_sugeridos.add("Elaboração do PGRS (Resíduos Perigosos)")
        servicos_sugeridos.add("Adequação da Caixa Separadora de Água e Óleo (SAO)")
    elif seg == "Alimentação (Restaurante/Mercado/Hotel)":
        servicos_sugeridos.add("Elaboração do PGRS (Alimentação)")
        servicos_sugeridos.add("Implantação de Coleta Seletiva e Compostagem")
    else:
        servicos_sugeridos.add("Elaboração do PGRS")
        servicos_sugeridos.add("Treinamento da Equipe Operacional")
        
    residuos = db["residuos_selecionados"]
    tem_perigoso = any("Grupo A" in r or "Grupo B" in r or "Óleo" in r or "Classe I" in r or "Classe D" in r or "Químico" in r for r in residuos)
    if tem_perigoso:
        servicos_sugeridos.add("Inventário Anual de Resíduos Perigosos (SINIR/SIGOR)")
        servicos_sugeridos.add("Emissão e Controle de MTR/CADRI")
        
    servicos_sugeridos.add("Contrato Mensal de Gestão e Acompanhamento Ambiental")
    return list(servicos_sugeridos)

if not db["servicos"]:
    db["servicos"] = gerar_servicos_recomendados()

# --- SIDEBAR (MENU LATERAL FIXO E INTERATIVO) ---
st.sidebar.markdown("<h2 style='color: white; font-weight: 900;'>🌱 EcoAudit Pro</h2>", unsafe_allow_html=True)
st.sidebar.markdown("<p style='color: #A7F3D0; font-size: 0.85rem;'>MENU DE NAVEGAÇÃO</p>", unsafe_allow_html=True)
st.sidebar.button("➕ Novo Relatório", on_click=resetar_dados, key="btn_novo_relatorio_sidebar", type="primary", use_container_width=True)
st.sidebar.markdown("---")

for item_aba in LISTA_ABAS:
    tipo_botao = "primary" if st.session_state.aba_ativa == item_aba else "secondary"
    if st.sidebar.button(item_aba, key=f"menu_{item_aba}", type=tipo_botao, use_container_width=True):
        st.session_state.aba_ativa = item_aba

aba_ativa = st.session_state.aba_ativa

# --- ESTÉTICA E CSS MOBILE-FIRST (CONTRASTE E CAIXAS CORRIGIDAS) ---
st.markdown("""
    <style>
    /* Fundo da aplicação */
    .stApp { 
        background-color: #F1F5F9 !important; 
        color: #0F172A !important;
    }

    /* Ajuste de espaçamento geral para telas móveis */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 2rem !important;
        padding-left: 0.8rem !important;
        padding-right: 0.8rem !important;
    }

    /* Menu Lateral */
    [data-testid="stSidebar"] {
        background-color: #064E3B !important; 
        padding-top: 10px;
    }
    
    [data-testid="stSidebar"] div.stButton > button {
        background: #047857 !important; 
        color: #FFFFFF !important; 
        font-size: 0.95rem !important;
        font-weight: 700 !important; 
        padding: 12px 16px !important; 
        border-radius: 12px !important;
        border: 1px solid #059669 !important; 
        text-align: left !important; 
        margin-bottom: 2px !important;
    }
    
    [data-testid="stSidebar"] div.stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        border: 2px solid #34D399 !important; 
        font-weight: 800 !important;
    }

    /* Cabeçalho Principal */
    .main-header {
        background: linear-gradient(135deg, #047857 0%, #065F46 50%, #064E3B 100%);
        padding: 16px 20px; 
        border-radius: 14px; 
        color: white; 
        margin-bottom: 16px;
        border-left: 6px solid #34D399;
    }
    .main-header h1 { 
        color: #FFFFFF !important; 
        font-weight: 900; 
        font-size: 1.5rem !important; 
        margin-bottom: 4px; 
    }
    .main-header p { 
        color: #A7F3D0 !important; 
        font-size: 0.85rem !important; 
        margin: 0; 
    }

    /* Cards de Conteúdo */
    .glass-card {
        background: #FFFFFF !important; 
        border-radius: 14px; 
        padding: 16px;
        border: 1px solid #CBD5E1; 
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        margin-bottom: 16px;
    }
    
    .card-title {
        color: #064E3B !important; 
        font-size: 1.1rem !important; 
        font-weight: 800 !important; 
        margin-bottom: 12px;
        border-bottom: 2px solid #A7F3D0; 
        padding-bottom: 6px;
    }

    /* CONTRASTE: CAIXAS DE ENTRADA, SELEÇÃO E RÓTULOS */
    .stTextInput label, .stNumberInput label, .stSelectbox label, 
    .stTextArea label, .stDateInput label, .stTimeInput label, 
    .stMultiSelect label, .stSlider label {
        color: #0F172A !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }

    div[data-baseweb="input"] {
        background-color: #FFFFFF !important;
        border: 1.5px solid #059669 !important;
        border-radius: 10px !important;
    }
    
    div[data-baseweb="input"] input {
        color: #0F172A !important;
        background-color: #FFFFFF !important;
        font-weight: 700 !important;
        -webkit-text-fill-color: #0F172A !important;
    }

    div[data-baseweb="textarea"] {
        background-color: #FFFFFF !important;
        border: 1.5px solid #059669 !important;
        border-radius: 10px !important;
    }

    div[data-baseweb="textarea"] textarea {
        color: #0F172A !important;
        background-color: #FFFFFF !important;
        font-weight: 600 !important;
        -webkit-text-fill-color: #0F172A !important;
    }

    div[data-baseweb="select"] > div {
        background-color: #FFFFFF !important;
        border: 1.5px solid #059669 !important;
        border-radius: 10px !important;
        color: #0F172A !important;
    }

    div[data-baseweb="select"] span {
        color: #0F172A !important;
        font-weight: 700 !important;
    }

    /* Botões operacionais */
    .stButton > button {
        width: 100% !important;
        border-radius: 10px !important;
        padding: 12px 16px !important;
        font-weight: 700 !important;
    }

    /* Indicadores / Cards Coloridos */
    .score-balloon { 
        background: #ECFDF5 !important; 
        border: 2px solid #10B981 !important; 
        border-radius: 14px; 
        padding: 16px; 
        text-align: center; 
        margin-bottom: 12px;
    }
    .score-number { 
        font-size: 2.8rem; 
        font-weight: 900; 
        color: #047857 !important; 
        margin: 6px 0; 
    }
    
    .multa-card { 
        background: #FEF2F2 !important; 
        border: 2px solid #F87171 !important; 
        border-radius: 14px; 
        padding: 16px; 
        text-align: center; 
        margin-bottom: 12px;
    }
    
    .economia-card { 
        background: #F0FDF4 !important; 
        border: 2px solid #4ADE80 !important; 
        border-radius: 14px; 
        padding: 16px; 
        text-align: center; 
        margin-bottom: 12px;
    }
    
    .badge-servico {
        background-color: #E0F2FE !important;
        border: 1px solid #0284C7 !important;
        color: #0369A1 !important;
        padding: 6px 10px;
        border-radius: 8px;
        font-weight: 700;
        font-size: 0.85rem;
        margin-bottom: 6px;
        display: inline-block;
    }
    </style>
""", unsafe_allow_html=True)

# --- CABEÇALHO ---
st.markdown("""
    <div class="main-header">
        <h1>🌱 EcoAudit Pro</h1>
        <p>Plataforma Profissional de Diagnóstico de Resíduos e Geração de PGRS</p>
    </div>
""", unsafe_allow_html=True)

# --- MENU DE NAVEGAÇÃO SUPERIOR & BOTÃO NOVO RELATÓRIO ---
col_nav1, col_nav2 = st.columns([3, 1])
with col_nav1:
    st.selectbox(
        "📍 Navegar pelas Páginas:",
        options=LISTA_ABAS,
        index=LISTA_ABAS.index(aba_ativa),
        key="w_select_navegacao",
        on_change=sincronizar_menu_dropdown
    )
with col_nav2:
    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
    st.button("➕ Novo Relatório", on_click=resetar_dados, key="btn_novo_relatorio_top", type="secondary", use_container_width=True)

# --- MÓDULO 1: LOCAL ---
if aba_ativa == "📋 1. Local":
    st.markdown('<div class="glass-card"><div class="card-title">👨‍💼 Informações Básicas do Cliente</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        db["consultor"] = st.text_input("❓ Quem é o Consultor Responsável?", value=db["consultor"], key="w_consultor")
        db["data_vistoria"] = st.date_input("❓ Qual a Data da Vistoria?", value=db["data_vistoria"], key="w_data")
        db["empresa"] = st.text_input("❓ Razão Social / Nome da Empresa?", value=db["empresa"], placeholder="Ex: Metalúrgica Silva", key="w_empresa")
        db["cnpj"] = st.text_input("❓ CNPJ / CPF da Empresa?", value=db["cnpj"], placeholder="00.000.000/0001-00", key="w_cnpj")
        
        idx_seg = NICHOS_DISPONIVEIS.index(db["segmento"]) if db["segmento"] in NICHOS_DISPONIVEIS else 5
        db["segmento"] = st.selectbox("🎯 Nicho / Segmento Operacional?", NICHOS_DISPONIVEIS, index=idx_seg, key="w_segmento")

    with col2:
        db["resp"] = st.text_input("❓ Responsável Acompanhando no Cliente?", value=db["resp"], key="w_resp")
        db["tel"] = st.text_input("❓ Telefone / WhatsApp de Contato?", value=db["tel"], key="w_tel")
        db["email"] = st.text_input("❓ E-mail Principal do Cliente?", value=db["email"], key="w_email")
        db["func"] = st.number_input("❓ Número Total de Funcionários?", min_value=1, value=db["func"], key="w_func")
        db["area"] = st.number_input("❓ Área Aproximada do Imóvel (m²)?", min_value=10.0, value=db["area"], step=50.0, key="w_area")

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="glass-card"><div class="card-title">⚙️ Rotina Operacional do Local</div>', unsafe_allow_html=True)
    st.write("**❓ Qual o Horário de Funcionamento?**")
    col_h1, col_h2 = st.columns(2)
    db["h_ini"] = col_h1.time_input("Horário Inicial", value=db["h_ini"], key="w_h_ini")
    db["h_fim"] = col_h2.time_input("Horário Final", value=db["h_fim"], key="w_h_fim")

    c_op2, c_op3 = st.columns(2)
    with c_op2:
        db["filiais"] = st.pills("🏢 Possui Filiais?", ["Não", "Sim"], default=db["filiais"], key="w_filiais")
        
    with c_op3:
        db["prod"] = st.pills("🏭 Processos Produtivos?", ["Não", "Sim"], default=db["prod"], key="w_prod")

    db["obs"] = st.text_area("❓ Observações Gerais da Operação:", value=db["obs"], key="w_obs")
    st.markdown('</div>', unsafe_allow_html=True)

    st.button("💾 Avançar para Mapeamento ➡️", on_click=pular_aba, args=(LISTA_ABAS[1],), type="primary")

# --- BASE DE DADOS ADAPTATIVA ---
residuos_por_nicho = {
    "Saúde / Veterinária": ["Biológicos/Infectantes (Grupo A)", "Químicos/Medicamentos (Grupo B)", "Rejeitos Comuns (Grupo D)", "Perfurocortantes (Grupo E)", "Lâmpadas/Pilhas", "Embalagens plásticas/papel"],
    "Oficina / Posto / Funilaria": ["Óleo Lubrificante Usado (OLUC)", "Filtros de Óleo", "Estopas/EPIs Contaminados", "Solventes/Tintas/Lixas", "Baterias automotivas", "Pneus", "Metais/Peças", "Plásticos/Papelão"],
    "Alimentação (Restaurante/Mercado/Hotel)": ["Orgânicos/Restos de Comida", "Óleo Vegetal Usado", "Embalagens de Plástico", "Papelão/Caixas", "Vidros/Garrafas", "Latas/Metais", "Rejeito Comum"],
    "Indústria PME": ["Sucata Metálica", "Sobra de Madeira/Pallets", "Lodo/Efluentes", "Embalagens Químicas", "Óleos Industriais", "Papel/Plástico Reciclável", "Rejeito de Processo"],
    "Construção Civil": ["Classe A (Concreto/Tijolo/Argamassa)", "Classe B (Plástico/Papel/Madeira/Metal)", "Classe C (Gesso)", "Classe D (Tintas/Solventes/Telhas Amianto)", "Entulho Misturado"],
    "Condomínio / Comercial": ["Papel/Papelão", "Plásticos Diversos", "Vidros", "Metais/Alumínio", "Orgânicos", "Rejeitos Domésticos", "Lâmpadas/Eletrônicos/Pilhas"]
}

perguntas_auditoria = {
    "Saúde / Veterinária": {"norma": "RDC Anvisa 222/18", "seg": "Segregação por grupo (A, B, D, E)?", "arm": "Abrigo com áreas limpa/suja separadas?", "dest": "Coleta por empresa especializada licenciada?"},
    "Oficina / Posto / Funilaria": {"norma": "NBR 10004 / ANP", "seg": "Óleos e estopas separados do lixo comum?", "arm": "Caixa separadora de água e óleo (SAO) operando?", "dest": "MTR de OLUC assinado com coletor ANP?"},
    "Alimentação (Restaurante/Mercado/Hotel)": {"norma": "Vigilância Sanitária Local", "seg": "Separação de orgânicos e óleo vegetal?", "arm": "Caixa de gordura com manutenção regular?", "dest": "Destinação do óleo para empresa cadastrada?"},
    "Indústria PME": {"norma": "PNRS / Licenciamento Ambiental", "seg": "Segregação de Classe I vs Classe II?", "arm": "Bacia de contenção na área de resíduos?", "dest": "MTR estadual e CDF emitidos?"},
    "Construção Civil": {"norma": "Resolução CONAMA 307/02", "seg": "Triagem das Classes A, B, C e D?", "arm": "Caçambas estáticas identificadas?", "dest": "Destinação apenas para ATT ou Aterro RCCN?"},
    "Condomínio / Comercial": {"norma": "Posturas Municipais", "seg": "Lixeiras de coleta seletiva padronizadas?", "arm": "Container com tampa e local coberto?", "dest": "Parceria com Coleta Seletiva/Cooperativa?"}
}

# --- MÓDULO 2: MAPEAMENTO ---
if aba_ativa == "♻️ 2. Mapeamento":
    st.markdown(f'<div class="glass-card"><div class="card-title">📦 Resíduos Típicos do Segmento: {db["segmento"]}</div>', unsafe_allow_html=True)
    residuos_sugeridos = residuos_por_nicho.get(db["segmento"], [])
    
    lista_temporaria_selecionados = []
    
    for idx, item in enumerate(residuos_sugeridos):
        ja_marcado = item in db["residuos_selecionados"]
        if st.checkbox(f"✔️ {item}", value=ja_marcado, key=f"w_res_{idx}"):
            lista_temporaria_selecionados.append(item)
            
    db["outros_residuos"] = st.text_input("➕ Adicionar outro resíduo específico não listado:", value=db["outros_residuos"], key="w_outros_res")
    if db["outros_residuos"]:
        lista_temporaria_selecionados.extend([x.strip() for x in db["outros_residuos"].split(",") if x.strip()])
    
    db["residuos_selecionados"] = lista_temporaria_selecionados
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="glass-card"><div class="card-title">⚖️ Métrica e Estimativa de Geração</div>', unsafe_allow_html=True)
    q1, q2 = st.columns(2)
    db["sabe_qtd"] = q1.pills("Sabe a quantidade mensal?", ["Não", "Sim"], default=db["sabe_qtd"], key="w_sabe_qtd")
    db["balanca"] = q2.pills("Possui controle/balança?", ["Não", "Sim"], default=db["balanca"], key="w_balanca")
    
    q3, q4 = st.columns(2)
    db["kg_dia"] = q3.number_input("Est. Estimada (kg/dia)", min_value=0.0, value=db["kg_dia"], key="w_kg_dia")
    db["kg_mes"] = q4.number_input("Est. Estimada (kg/mês)", min_value=0.0, value=db["kg_mes"], key="w_kg_mes")
    st.markdown('</div>', unsafe_allow_html=True)

    st.button("💾 Avançar para Auditoria ➡️", on_click=pular_aba, args=(LISTA_ABAS[2],), type="primary")

# --- MÓDULO 3: AUDITORIA OPERACIONAL ---
if aba_ativa == "🔎 3. Auditoria":
    cfg_nicho = perguntas_auditoria.get(db["segmento"], perguntas_auditoria["Condomínio / Comercial"])
    st.info(f"💡 Auditoria adaptada para as normas de **{db['segmento']}** ({cfg_nicho['norma']}).")

    def render_eval_block(title, options_dict, start_idx):
        st.markdown(f'<div class="glass-card"><div class="card-title">{title}</div>', unsafe_allow_html=True)
        score = 0
        max_score = len(options_dict) * 10
        
        for i, (k, label) in enumerate(options_dict.items()):
            val_salvo = db["aud_resp"].get(k, "❌ Não")
            val = st.segmented_control(label, ["❌ Não", "⚠️ Parcial", "✅ Sim"], default=val_salvo, key=f"w_aud_{k}")
            db["aud_resp"][k] = val
            if val == "✅ Sim": score += 10
            elif val == "⚠️ Parcial": score += 5
            
        st.markdown('</div>', unsafe_allow_html=True)
        return round((score / max_score) * 10, 1) if max_score > 0 else 0

    n1 = render_eval_block("1. 🚮 Segregação e Práticas", {"s1": cfg_nicho["seg"], "s2": "Equipe capacitada?", "s3": "Procedimento interno?"}, 1)
    n2 = render_eval_block("2. 🗑️ Lixeiras e Coletores", {"l1": "Qtd. suficiente?", "l2": "Identificadas?", "l3": "Tampa/pedal?", "l4": "Cores padrão?"}, 2)
    n3 = render_eval_block("3. 🚚 Coleta Interna", {"c1": "Coleta em rotina?", "c2": "Carrinho próprio?", "c3": "Livre contaminação?"}, 3)
    n4 = render_eval_block("4. 🏭 Armazenamento Temporário", {"a1": cfg_nicho["arm"], "a2": "Coberta/Impermeável?", "a3": "Acesso controlado?", "a4": "Sinalizada?"}, 4)
    n5 = render_eval_block("5. 📜 Coleta Externa & Destinação", {"d1": cfg_nicho["dest"], "d2": "Existe contrato?", "d3": "Destino conhecido?", "d4": "MTR emitido?"}, 5)
    n6 = render_eval_block("6. 📄 Documentação e Licenças", {"doc1": "Possui PGRS/PGRSS?", "doc2": "Atualizado?", "doc3": "Inventário ok?", "doc4": "Licença válida?"}, 6)
    n7 = render_eval_block("7. 🎓 Treinamentos e Capacitação", {"t1": "Treinado 12m?", "t2": "Lista de presença?", "t3": "Material visível?"}, 7)
    n8 = render_eval_block("8. 🦺 Segurança e EPIs", {"sg1": "Uso EPIs ok?", "sg2": "Segurança sanitária ok?", "sg3": "Kits de vazamento ok?"}, 8)

    db["notas"] = [n1, n2, n3, n4, n5, n6, n7, n8]

    st.button("💾 Avançar para Fotos/Riscos ➡️", on_click=pular_aba, args=(LISTA_ABAS[3],), type="primary")

# --- MÓDULO 4: CÂMERA & RISCOS ---
if aba_ativa == "📸 4. Fotos & Riscos":
    st.markdown('<div class="glass-card"><div class="card-title">📸 Registro Fotográfico em Tempo Real</div>', unsafe_allow_html=True)
    foto_tirada = st.camera_input("Capturar foto do local", key="w_cam")
    if foto_tirada:
        db["foto_memoria"] = foto_tirada
        st.success("📸 Foto capturada e anexada ao Relatório!")
    
    st.file_uploader("Selecione fotos salvas da galeria", type=["jpg", "png", "jpeg"], accept_multiple_files=True, key="w_galeria")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="glass-card"><div class="card-title">⚠️ Matriz de Riscos Detectados</div>', unsafe_allow_html=True)
    db["rk_legal"] = st.select_slider("⚖️ Risco Legal", ["Baixo", "Médio", "Alto"], value=db["rk_legal"], key="w_rk_legal")
    db["rk_amb"] = st.select_slider("🍃 Risco Ambiental", ["Baixo", "Médio", "Alto"], value=db["rk_amb"], key="w_rk_amb")
    db["rk_op"] = st.select_slider("⚙️ Risco Operacional", ["Baixo", "Médio", "Alto"], value=db["rk_op"], key="w_rk_op")
    
    db["nc"] = st.text_area("Descreva as Não Conformidades Encontradas:", value=db["nc"], key="w_nc")
    db["rec"] = st.text_area("💡 Recomendações Imediatas:", value=db["rec"], key="w_rec")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="glass-card"><div class="card-title">💼 Oportunidades Comerciais & Serviços Recomendados</div>', unsafe_allow_html=True)
    
    sugestoes_auto = gerar_servicos_recomendados()
    st.markdown("**🤖 Serviços Sugeridos pelo Diagnóstico:**")
    for s_sug in sugestoes_auto:
        st.markdown(f'<span class="badge-servico">✨ {s_sug}</span>', unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    opcoes_completas = sorted(list(set(sugestoes_auto + [
        "Elaboração do PGRS", "Elaboração do PGRS/PGRSS", "Elaboração do PGRSS (Saúde)", 
        "Elaboração do PGRCC (Obras/Construção)", "Elaboração do PGRS (Resíduos Perigosos)", 
        "Elaboração do PGRS (Alimentação)", "Treinamento de Biossegurança e Perfurocortantes", 
        "Homologação de Destinadores e ATT/Aterros", "Adequação da Caixa Separadora de Água e Óleo (SAO)", 
        "Implantação de Coleta Seletiva e Compostagem", "Inventário Anual de Resíduos Perigosos (SINIR/SIGOR)", 
        "Emissão e Controle de MTR/CADRI", "Treinamento da Equipe Operacional", 
        "Contrato Mensal de Gestão e Acompanhamento Ambiental"
    ])))
    
    valores_salvos = db["servicos"] if db["servicos"] else sugestoes_auto
    valores_validos = [s for s in valores_salvos if s in opcoes_completas]
    
    db["servicos"] = st.multiselect(
        "Selecione os serviços que farão parte da Proposta Comercial:",
        options=opcoes_completas,
        default=valores_validos,
        key="w_servicos"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.button("💾 Avançar para Score/Multas ➡️", on_click=pular_aba, args=(LISTA_ABAS[4],), type="primary")

# --- MÓDULO 5: SCORE & MULTAS ---
if aba_ativa == "📊 5. Score & Multas":
    nota_final = round((sum(db["notas"]) / 80) * 100, 1) if sum(db["notas"]) > 0 else 0.0

    if nota_final >= 90: situacao, cor_badge, fator_multa = "Excelente 🎉", "#059669", 0.0
    elif nota_final >= 70: situacao, cor_badge, fator_multa = "Boa 👍", "#0284C7", 0.15
    elif nota_final >= 50: situacao, cor_badge, fator_multa = "Regular ⚠️", "#D97706", 0.40
    elif nota_final >= 30: situacao, cor_badge, fator_multa = "Crítica 🚨", "#DC2626", 0.75
    else: situacao, cor_badge, fator_multa = "Muito Crítica 💀", "#7F1D1D", 1.0

    base_multa_segmento = {
        "Saúde / Veterinária": 15000.0, "Oficina / Posto / Funilaria": 20000.0,
        "Alimentação (Restaurante/Mercado/Hotel)": 10000.0, "Indústria PME": 35000.0,
        "Construção Civil": 25000.0, "Condomínio / Comercial": 8000.0
    }
    multa_estimada_min = round(base_multa_segmento.get(db["segmento"], 12000.0) * fator_multa, 2)
    multa_estimada_max = round(multa_estimada_min * 3.5, 2)
    economia_ano = round(((db["kg_mes"] if db["kg_mes"] > 0 else 300.0) * 1.50 * 0.35) * 12, 2)

    st.markdown(f"""
        <div class="score-balloon">
            <span style="font-weight:800; color:#064E3B; text-transform:uppercase;">Score do Nicho: {db["segmento"]}</span>
            <div class="score-number">{nota_final}<span style="font-size:1.5rem">/100</span></div>
            <div style="font-size:1.3rem; font-weight:800; color:{cor_badge};">{situacao}</div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="glass-card"><div class="card-title">📌 Desempenho Técnico das Áreas</div>', unsafe_allow_html=True)
    df_notas = pd.DataFrame({"Área Avaliada": ["Segregação", "Lixeiras", "Coleta Int.", "Armazenamento", "Destinação", "Documentação", "Treinamento", "Segurança"], "Nota (0-10)": db["notas"]})
    st.dataframe(df_notas, column_config={"Nota (0-10)": st.column_config.ProgressColumn("Desempenho", format="%.1f pts", min_value=0, max_value=10)}, use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(f"""
        <div class="multa-card">
            <h4 style="color:#991B1B; margin:0;">🚨 Risco de Multa Estimado</h4>
            <div style="font-size:1.8rem; font-weight:900; color:#DC2626; margin:8px 0;">R$ {multa_estimada_min:,.2f} a R$ {multa_estimada_max:,.2f}</div>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown(f"""
        <div class="economia-card">
            <h4 style="color:#065F46; margin:0;">💡 Economia Anual Estimada</h4>
            <div style="font-size:1.8rem; font-weight:900; color:#059669; margin:8px 0;">R$ {(economia_ano + multa_estimada_min):,.2f} /ano</div>
        </div>
    """, unsafe_allow_html=True)

    st.button("💾 Avançar para Documentos ➡️", on_click=pular_aba, args=(LISTA_ABAS[5],), type="primary")

# --- MÓDULO 6: DOCUMENTOS ---
if aba_ativa == "📄 6. Documentos":
    st.markdown(f'<div class="glass-card"><div class="card-title">📄 Documentos Obrigatórios do Segmento: {db["segmento"]}</div>', unsafe_allow_html=True)
    st.markdown("#### 📜 Documentos Gerais da Empresa")
    
    def salvar_doc(chave, titulo):
        db["docs"][chave] = st.checkbox(titulo, value=db["docs"].get(chave, False), key=f"w_doc_{chave}")
        
    salvar_doc("cnpj", "📄 Cartão CNPJ / Contrato Social")
    salvar_doc("alvara", "🏢 Alvará de Funcionamento")
    salvar_doc("avcb", "🧯 AVCB / CLCB do Corpo de Bombeiros")
    salvar_doc("art", "🖊️ ART/CFT de Responsabilidade Técnica")

    st.markdown("---")
    st.markdown(f"#### 🎯 Documentos Específicos para {db['segmento']}")
    if db["segmento"] == "Saúde / Veterinária":
        salvar_doc("saude1", "🩺 Licença Sanitária (LVS/CVS)")
        salvar_doc("saude2", "🚛 Contrato com Coletora Infectantes (Grupo A/E)")
        salvar_doc("saude3", "📦 MTRs e CDFs de Resíduos de Saúde")
    elif db["segmento"] == "Oficina / Posto / Funilaria":
        salvar_doc("ofi1", "📜 Licença Ambiental de Operação (LO)")
        salvar_doc("ofi2", "🛢️ Contrato e MTR de OLUC (Óleo Usado - ANP)")
    else:
        salvar_doc("gen1", "📜 Comprovante de Coleta Seletiva / Cooperativa")

    st.markdown('</div>', unsafe_allow_html=True)

    st.button("💾 Avançar para Plano/PGRS ➡️", on_click=pular_aba, args=(LISTA_ABAS[6],), type="primary")

# --- MÉTODOS DE EXPORTAÇÃO (PDF E EXCEL) ---
def limpar_texto_pdf(texto):
    return re.sub(r'[^\x00-\xFF]', '', str(texto))

def gerar_pdf_relatorio():
    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_fill_color(5, 150, 105)
    pdf.rect(0, 0, 210, 32, 'F')
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, limpar_texto_pdf("RELATÓRIO DE DIAGNÓSTICO AMBIENTAL"), ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, limpar_texto_pdf(f"Cliente: {db['empresa']} | Nicho: {db['segmento']}"), ln=True, align="C")
    pdf.ln(12)
    
    pdf.set_text_color(15, 23, 42)
    
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(226, 232, 240)
    pdf.cell(0, 7, limpar_texto_pdf(" 1. DADOS DO CLIENTE E VISTORIA"), ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.ln(2)
    
    pdf.cell(100, 5, limpar_texto_pdf(f"Razão Social: {db['empresa']}"), ln=0)
    pdf.cell(90, 5, limpar_texto_pdf(f"CNPJ: {db['cnpj']}"), ln=1)
    
    pdf.cell(100, 5, limpar_texto_pdf(f"Responsável: {db['resp']}"), ln=0)
    pdf.cell(90, 5, limpar_texto_pdf(f"Contato: {db['tel']} | {db['email']}"), ln=1)
    
    pdf.cell(100, 5, limpar_texto_pdf(f"Data Vistoria: {db['data_vistoria'].strftime('%d/%m/%Y')}"), ln=0)
    pdf.cell(90, 5, limpar_texto_pdf(f"Consultor Técnico: {db['consultor']}"), ln=1)
    
    pdf.cell(100, 5, limpar_texto_pdf(f"Nº Funcionários: {db['func']} colaboradores"), ln=0)
    pdf.cell(90, 5, limpar_texto_pdf(f"Área Construída: {db['area']} m²"), ln=1)
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, limpar_texto_pdf(" 2. INDICADORES DE RISCO E DIAGNÓSTICO FINANCEIRO"), ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.ln(2)

    nota_final = round((sum(db["notas"]) / 80) * 100, 1) if sum(db["notas"]) > 0 else 0.0
    if nota_final >= 90: situacao, fator_multa = "Excelente", 0.0
    elif nota_final >= 70: situacao, fator_multa = "Boa", 0.15
    elif nota_final >= 50: situacao, fator_multa = "Regular", 0.40
    elif nota_final >= 30: situacao, fator_multa = "Crítica", 0.75
    else: situacao, fator_multa = "Muito Crítica", 1.0

    base_multa_segmento = {
        "Saúde / Veterinária": 15000.0, "Oficina / Posto / Funilaria": 20000.0,
        "Alimentação (Restaurante/Mercado/Hotel)": 10000.0, "Indústria PME": 35000.0,
        "Construção Civil": 25000.0, "Condomínio / Comercial": 8000.0
    }
    multa_min = round(base_multa_segmento.get(db["segmento"], 12000.0) * fator_multa, 2)
    multa_max = round(multa_min * 3.5, 2)
    economia_ano = round(((db["kg_mes"] if db["kg_mes"] > 0 else 300.0) * 1.50 * 0.35) * 12, 2)

    pdf.cell(100, 5, limpar_texto_pdf(f"Score do Nicho: {nota_final} / 100 ({situacao})"), ln=0)
    pdf.cell(90, 5, limpar_texto_pdf(f"Economia Anual Estimada: R$ {(economia_ano + multa_min):,.2f}"), ln=1)
    
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(100, 5, limpar_texto_pdf(f"RISCO DE MULTA ESTIMADO: R$ {multa_min:,.2f} a R$ {multa_max:,.2f}"), ln=1)
    
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(63, 5, limpar_texto_pdf(f"Risco Legal: {db['rk_legal']}"), ln=0)
    pdf.cell(63, 5, limpar_texto_pdf(f"Risco Ambiental: {db['rk_amb']}"), ln=0)
    pdf.cell(63, 5, limpar_texto_pdf(f"Risco Operacional: {db['rk_op']}"), ln=1)
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, limpar_texto_pdf(" 3. MAPEAMENTO DE RESÍDUOS GERADOS"), ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.ln(2)
    
    res_str = ", ".join(db["residuos_selecionados"]) if db["residuos_selecionados"] else "Nenhum resíduo mapeado."
    pdf.multi_cell(0, 5, limpar_texto_pdf(f"Resíduos Identificados: {res_str}"))
    pdf.cell(0, 5, limpar_texto_pdf(f"Geração Estimada: ~{db['kg_dia']} kg/dia (Aprox. {db['kg_mes']} kg/mês)"), ln=True)
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, limpar_texto_pdf(" 4. NÃO CONFORMIDADES E RECOMENDAÇÕES"), ln=True, fill=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.ln(2)
    
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, limpar_texto_pdf("Inconformidades Identificadas:"), ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, limpar_texto_pdf(db["nc"] if db["nc"] else "Nenhuma inconformidade registrada."))
    pdf.ln(2)
    
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, limpar_texto_pdf("Recomendações Técnicas Imediatas:"), ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, limpar_texto_pdf(db["rec"] if db["rec"] else "Sem recomendações cadastradas."))
    pdf.ln(5)

    if db.get("servicos"):
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, limpar_texto_pdf(" 5. PLANO DE ADEQUAÇÃO E SERVIÇOS RECOMENDADOS"), ln=True, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.ln(2)
        for srv in db["servicos"]:
            pdf.cell(0, 5, limpar_texto_pdf(f"• {srv}"), ln=True)
        pdf.ln(5)

    if db.get("foto_memoria") is not None:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_file:
                tmp_file.write(db["foto_memoria"].getvalue())
                tmp_path = tmp_file.name

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 7, limpar_texto_pdf(" 6. REGISTRO FOTOGRÁFICO DA VISTORIA"), ln=True, fill=True)
            pdf.ln(3)
            pdf.image(tmp_path, x=10, w=90)
        except Exception:
            pass

    pdf_bytes = pdf.output()
    if type(pdf_bytes) == str:
        pdf_bytes = pdf_bytes.encode('latin1')
    return bytes(pdf_bytes)

def gerar_pdf_minuta_pgrs():
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.add_page()
    pdf.set_fill_color(6, 78, 59)
    pdf.rect(0, 0, 210, 297, 'F')
    
    nome_doc = "PGRSS" if db["segmento"] == "Saúde / Veterinária" else ("PGRCC" if db["segmento"] == "Construção Civil" else "PGRS")
    
    pdf.ln(35)
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, limpar_texto_pdf("PLANO DE GERENCIAMENTO DE"), ln=True, align="C")
    pdf.cell(0, 12, limpar_texto_pdf(f"RESÍDUOS SÓLIDOS ({nome_doc})"), ln=True, align="C")
    pdf.ln(8)
    
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, limpar_texto_pdf(f"Minuta Técnica Oficial - Segmento: {db['segmento']}"), ln=True, align="C")
    
    pdf.ln(70)
    pdf.set_font("Helvetica", "B", 15)
    emp_titulo = db['empresa'].upper() if db['empresa'] else "[RAZÃO SOCIAL DO CLIENTE]"
    pdf.cell(0, 8, limpar_texto_pdf(f"EMPRESA: {emp_titulo}"), ln=True, align="C")
    
    pdf.set_font("Helvetica", "", 11)
    cnpj_txt = db['cnpj'] if db['cnpj'] else "[00.000.000/0001-00]"
    pdf.cell(0, 6, limpar_texto_pdf(f"CNPJ / CPF: {cnpj_txt}"), ln=True, align="C")
    
    pdf.ln(50)
    pdf.cell(0, 6, limpar_texto_pdf(f"Responsável Técnico: {db['consultor']}"), ln=True, align="C")
    pdf.cell(0, 6, limpar_texto_pdf(f"Data de Emissão: {db['data_vistoria'].strftime('%d/%m/%Y')}"), ln=True, align="C")

    pdf.add_page()
    pdf.set_text_color(15, 23, 42)
    
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(226, 232, 240)
    pdf.cell(0, 8, limpar_texto_pdf(" 1. IDENTIFICAÇÃO DO EMPREENDIMENTO E RESPONSÁVEIS"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(100, 6, limpar_texto_pdf(f"Razão Social: {db['empresa'] if db['empresa'] else '[Preencher]'}" ), ln=0)
    pdf.cell(90, 6, limpar_texto_pdf(f"CNPJ: {db['cnpj'] if db['cnpj'] else '[Preencher]'}" ), ln=1)
    
    pdf.cell(100, 6, limpar_texto_pdf(f"Segmento Operacional: {db['segmento']}"), ln=0)
    pdf.cell(90, 6, limpar_texto_pdf(f"Área Construída/Útil: {db['area']} m²"), ln=1)
    
    pdf.cell(100, 6, limpar_texto_pdf(f"Nº de Colaboradores: {db['func']} pessoas"), ln=0)
    pdf.cell(90, 6, limpar_texto_pdf(f"Horário: {db['h_ini'].strftime('%H:%M')} às {db['h_fim'].strftime('%H:%M')}"), ln=1)
    
    pdf.cell(100, 6, limpar_texto_pdf(f"Responsável Interno: {db['resp'] if db['resp'] else '[Preencher]'}" ), ln=0)
    pdf.cell(90, 6, limpar_texto_pdf(f"Contato: {db['tel']} | {db['email']}"), ln=1)
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpar_texto_pdf(" 2. EMBASAMENTO LEGAL E OBJETIVOS DO PLANO"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, limpar_texto_pdf(
        "Este Plano de Gerenciamento de Resíduos Sólidos tem como objetivo estabelecer as diretrizes socioambientais "
        "e operacionais para o manejo correto, minimização da geração, segregação, acondicionamento, armazenamento, "
        "transporte e destinação final ambientalmente adequada de todos os resíduos gerados nas atividades da empresa.\n\n"
        "O presente documento fundamenta-se tecnicamente e juridicamente nas seguintes normas vigentes:\n"
        "• Lei Federal nº 12.305/2010 - Política Nacional de Resíduos Sólidos (PNRS);\n"
        "• ABNT NBR 10004:2004 - Resíduos Sólidos - Classificação;\n"
        "• Diretrizes vigentes do Órgão Ambiental Estadual e Posturas Municipais aplicáveis."
    ))
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpar_texto_pdf(" 3. CARACTERIZAÇÃO E QUANTIFICAÇÃO DOS RESÍDUOS"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    res_lista = db['residuos_selecionados'] if db['residuos_selecionados'] else ['[Nenhum resíduo mapeado]']
    res_txt = ", ".join(res_lista)
    
    pdf.multi_cell(0, 5, limpar_texto_pdf(
        f"Com base na vistoria técnica de campo, foram mapeadas as seguintes tipologias de resíduos:\n"
        f"• Tipos de Resíduos Mapeados: {res_txt}\n"
        f"• Estimativa de Geração Diária: ~{db['kg_dia']} kg/dia\n"
        f"• Estimativa de Geração Mensal: ~{db['kg_mes']} kg/mês\n\n"
        f"Observações Operacionais: {db['obs'] if db['obs'] else 'Operação dentro da rotina padrão.'}"
    ))
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpar_texto_pdf(" 4. PROCEDIMENTOS OPERACIONAIS DE MANEJO INTERNO E EXTERNO"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, limpar_texto_pdf(
        "4.1. Segregação e Acondicionamento na Fonte:\n"
        "A segregação dos resíduos deve ser realizada diretamente na fonte geradora, utilizando coletores "
        "identificados com cores padronizadas e recipientes adequados ao tipo e classe do resíduo.\n\n"
        "4.2. Coleta e Transporte Interno:\n"
        "A coleta interna ocorrerá diariamente em rotinas definidas, utilizando equipamentos adequados de transporte "
        "para garantir a higiene e evitar contaminação cruzada ou riscos ocupacionais.\n\n"
        "4.3. Armazenamento Temporário:\n"
        "Os resíduos recolhidos serão mantidos no abrigo temporário de resíduos até o recolhimento pela coleta pública "
        "ou empresa terceirizada especializada, devendo a área ser coberta, ventilada e com acesso restrito.\n\n"
        "4.4. Destinação Final e Rastreabilidade Documental:\n"
        "Todos os resíduos recicláveis e perigosos serão destinados exclusivamente a empresas receptoras devidamente licenciadas. "
        "A rastreabilidade é garantida pela emissão contínua do Manifesto de Transporte de Resíduos (MTR) e arquivamento dos Certificados de Destinação Final (CDF)."
    ))
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpar_texto_pdf(" 5. PLANO DE AÇÃO E CRONOGRAMA DE ADEQUAÇÃO"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, limpar_texto_pdf(
        f"Ações Imediatas Prioritárias:\n"
        f"1. Não Conformidades a Corrigir: {db['nc']}\n\n"
        f"2. Recomendações Técnicas Imediatas: {db['rec']}\n\n"
        f"3. Capacitação e Treinamentos: Realizar treinamentos com a equipe operacional em até 30 dias para garantir o cumprimento integral deste plano."
    ))
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpar_texto_pdf(" 6. TERMO DE COMPROMISSO E ENCERRAMENTO"), ln=True, fill=True)
    pdf.ln(3)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, limpar_texto_pdf(
        "Declaramos que as informações contidas nesta minuta refletem fielmente a estrutura operacional e a rotina do estabelecimento, "
        "comprometendo-se as partes com a implantação e manutenção contínua das diretrizes estabelecidas."
    ))
    pdf.ln(25)
    
    pdf.cell(95, 6, "___________________________________", ln=0, align="C")
    pdf.cell(95, 6, "___________________________________", ln=1, align="C")
    pdf.cell(95, 5, limpar_texto_pdf(f"{db['empresa'] if db['empresa'] else 'Representante Legal'}"), ln=0, align="C")
    pdf.cell(95, 5, limpar_texto_pdf(f"{db['consultor']} - Responsável Técnico (ART)"), ln=1, align="C")

    pdf_bytes = pdf.output()
    if type(pdf_bytes) == str:
        pdf_bytes = pdf_bytes.encode('latin1')
    return bytes(pdf_bytes)

def gerar_excel_avancado():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Executivo"
    
    green_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    ws.merge_cells("A1:E1")
    ws["A1"] = f"DIAGNÓSTICO AMBIENTAL & AUDITORIA - {db['empresa'].upper()}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="064E3B")
    
    dados = [
        ["Consultor:", db["consultor"], "", "Data:", db["data_vistoria"].strftime("%d/%m/%Y")],
        ["Empresa:", db["empresa"], "", "CNPJ:", db["cnpj"]],
        ["Segmento:", db["segmento"], "", "Funcionários:", db["func"]]
    ]
    for row_idx, row_data in enumerate(dados, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws_plano = wb.create_sheet(title="Plano de Acao")
    ws_plano.append(["Prioridade", "O que fazer (What)", "Por que fazer (Why)", "Prazo"])
    for col_num in range(1, 5):
        c = ws_plano.cell(row=1, column=col_num)
        c.font = header_font
        c.fill = green_header_fill
        
    ws_plano.append(["Alta", "Elaborar / Atualizar o PGRS", "Atender legislação vigente", "30 dias"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- MÓDULO 7: PLANO DE AÇÃO & MINUTA PGRS ---
if aba_ativa == "🚀 7. Plano de Ação & Minuta PGRS":
    st.markdown(f'<div class="glass-card"><div class="card-title">🚀 Plano de Ação Personalizado para {db["segmento"]}</div>', unsafe_allow_html=True)
    acoes_sugeridas = [{"Prioridade": "🔥 Alta", "O que fazer": "Elaborar / Atualizar o PGRS", "Por que": "Atender legislação", "Prazo": "30 dias"}]
    st.dataframe(pd.DataFrame(acoes_sugeridas), use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    nome_base = formatar_nome_arquivo(db["empresa"] if db["empresa"] else "Empresa", db["segmento"])

    st.markdown('<div class="glass-card"><div class="card-title">📥 Gerar Minuta do PGRS & Relatórios Técnicos</div>', unsafe_allow_html=True)
    
    st.markdown("##### 📄 Relatório Técnico")
    st.download_button("🔴 BAIXAR RELATÓRIO", data=gerar_pdf_relatorio(), file_name=f"Relatorio_{nome_base}.pdf", mime="application/pdf", use_container_width=True)

    st.markdown("##### 📜 Minuta do PGRS")
    st.download_button("📘 BAIXAR MINUTA PGRS", data=gerar_pdf_minuta_pgrs(), file_name=f"Minuta_{nome_base}.pdf", mime="application/pdf", use_container_width=True)

    st.markdown("##### 📊 Planilha Excel")
    st.download_button("🟢 BAIXAR EXCEL", data=gerar_excel_avancado(), file_name=f"Auditoria_{nome_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.button("💾 Avançar para Calculadora ➡️", on_click=pular_aba, args=(LISTA_ABAS[7],), type="primary")

# --- MÓDULO 8: CALCULADORA DE ORÇAMENTO ---
if aba_ativa == "💰 8. Calculadora de Orçamento":
    st.markdown('<div class="glass-card"><div class="card-title">💰 Precificação Automática Dinâmica</div>', unsafe_allow_html=True)
    
    if not db["servicos"]:
        db["servicos"] = gerar_servicos_recomendados()

    tabela_horas = {
        "PGRS": 8, "PGRSS": 10, "PGRCC": 10, "Treinamento": 4, "Inventário": 5,
        "MTR": 3, "CADRI": 4, "Coleta Seletiva": 6, "SAO": 5, "Gestão": 6
    }
    
    detalhes_servicos = []
    horas_totais_servicos = 0
    
    for srv in db["servicos"]:
        h_srv = 5
        for kw, h_val in tabela_horas.items():
            if kw.lower() in srv.lower():
                h_srv = h_val
                break
        horas_totais_servicos += h_srv
        detalhes_servicos.append({"Serviço Selecionado": srv, "Horas Estimadas": f"{h_srv}h"})

    st.markdown("#### 📋 Serviços Incluídos na Proposta:")
    df_srv = pd.DataFrame(detalhes_servicos)
    st.dataframe(df_srv, use_container_width=True, hide_index=True)
    
    h_visita_base = 4
    h_area = int(db["area"] / 250)
    horas_calculadas = h_visita_base + h_area + horas_totais_servicos

    st.markdown("---")
    st.markdown(f"**⏱️ Horas Calculadas Automaticamente:** `{h_visita_base}h (Visita)` + `{h_area}h (Área)` + `{horas_totais_servicos}h (Serviços)` = **`{horas_calculadas} horas`**")
    
    horas_finais = st.number_input("⏱️ Total de Horas Técnicas Utilizadas no Cálculo:", min_value=1, value=horas_calculadas, key="w_h_calc")
    db["v_hora"] = st.number_input("❓ Valor da Sua Hora Técnica (R$):", min_value=10.0, value=db["v_hora"], step=10.0, key="w_v_hora")
    db["c_extra"] = st.number_input("❓ Custos Extras (Deslocamento/ART):", min_value=0.0, value=db["c_extra"], step=20.0, key="w_c_extra")
    db["margem"] = st.slider("❓ Margem de Lucro (%):", 0, 100, value=db["margem"], key="w_margem")

    custo_base = (horas_finais * db["v_hora"]) + db["c_extra"]
    valor_proposta = custo_base * (1 + (db["margem"] / 100))

    st.markdown(f'<div class="score-balloon"><span style="font-weight:800;">CUSTO BRUTO DE EXECUÇÃO</span><div style="font-size:1.8rem; font-weight:900;">R$ {custo_base:,.2f}</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="score-balloon"><span style="font-weight:800; color:#064E3B;">VALOR FINAL RECOMENDADO</span><div style="font-size:2rem; font-weight:900; color:#059669;">R$ {valor_proposta:,.2f}</div></div>', unsafe_allow_html=True)

    st.button("💾 Avançar para Passo a Passo ➡️", on_click=pular_aba, args=(LISTA_ABAS[8],), type="primary")

# --- MÓDULO 9: PASSO A PASSO ---
if aba_ativa == "📋 9. Passo a Passo do Gerenciamento":
    st.markdown('<div class="glass-card"><div class="card-title">📚 Tutorial Prático de Elaboração do PGRS</div>', unsafe_allow_html=True)
    st.markdown("""
    Este guia serve como um roteiro estruturado para conduzir a consultoria técnica e elaborar o Plano de Gerenciamento de Resíduos Sólidos (PGRS) em conformidade com a **Lei Federal nº 12.305/10 (PNRS)**.

    ---

    #### **Passo 1: Caracterização e Dados do Gerador**
    * **Identificação Completa:** Razão Social, CNPJ, Nome Fantasia, Endereço, Inscrição Estadual e Municipal.
    * **Dados Operacionais:** Área total (m²), número de colaboradores, horário de funcionamento e capacidade produtiva/atendimento.
    * **Responsáveis:** Identificar o Responsável Legal pelo estabelecimento e o Responsável Técnico pela elaboração/implantação do PGRS.

    #### **Passo 2: Mapeamento de Pontos Geradores e Gravimetria**
    * **Mapeamento de Processos:** Mapear cada setor do empreendimento (administração, produção, refeitório, manutenção).
    * **Classificação Normativa:**
      * **Indústrias/Gerais:** NBR 10004 (Classe I - Perigosos; Classe II A - Não Inertes; Classe II B - Inertes).
      * **Saúde:** RDC Anvisa 222/18 (Grupos A, B, C, D e E).
      * **Construção Civil:** Resolução CONAMA 307/02 (Classes A, B, C e D).
    * **Estimativa de Volume:** Quantificar a geração diária e mensal.

    #### **Passo 3: Mapeamento do Manejo Interno**
    * **Segregação na Fonte:** Estabelecer a separação no momento exato da geração.
    * **Acondicionamento:** Lixeiras com pedal, sacos coloridos padronizados, coletores de perfurocortantes, etc.
    * **Coleta e Transporte Interno:** Definir rotas internas e horários de recolhimento para evitar contaminação cruzada.

    #### **Passo 4: Estruturação do Abrigo Temporário de Resíduos**
    * **Infraestrutura Física:** Área com piso impermeável, paredes laváveis, cobertura contra intempéries e ventilação adequada.
    * **Segurança e Higiene:** Ponto de água para lavagem, canaleta de contenção/drenagem, acesso restrito.

    #### **Passo 5: Logística Externa e Homologação de Fornecedores**
    * **Transportadores e Destinadores:** Cadastrar empresas terceirizadas licenciadas pelo Órgão Ambiental.
    * **Rastreabilidade Documental:** Exigir o **MTR** e arquivar os **CDFs**. Manter atualizado o Inventário Nacional (SINIR/SIGOR).

    #### **Passo 6: Capacitação, Metas e Acompanhamento**
    * **Plano de Treinamento:** Treinamentos admissionais e periódicos com lista de presença.
    * **Responsabilidade Técnica:** Anexar a **ART** devidamente quitada.
    """)
    st.button("💾 Avançar para Legislação ➡️", on_click=pular_aba, args=(LISTA_ABAS[9],), type="primary")

# --- MÓDULO 10: LEGISLAÇÃO ---
if aba_ativa == "📜 10. Legislação & Sistemas":
    st.markdown('<div class="glass-card"><div class="card-title">📖 Guia Técnico de Legislação e Sistemas Regulatórios</div>', unsafe_allow_html=True)
    st.caption("Consulte as principais diretrizes ambientais para fundamentar pareceres e relatórios técnicos:")

    st.markdown("### 🏛️ **Principais Leis e Normas Federais**")
    st.markdown("""
    * **Lei Federal nº 12.305/2010 (PNRS)** [🔗](https://www.planalto.gov.br/ccivil_03/_ato2007-2010/2010/lei/l12305.htm): Institui a Política Nacional de Resíduos Sólidos, responsabilidade compartilhada e a obrigatoriedade do PGRS.
    * **ABNT NBR 10004:2004** [🔗](https://cetesb.sp.gov.br/wp-content/uploads/2014/12/NBR10004.pdf):
      * **Classe I (Perigosos):** Óleos, filtros, químicos (inflamáveis, tóxicos).
      * **Classe II A (Não Inertes):** Papel, plástico, orgânicos.
      * **Classe II B (Inertes):** Entulho limpo, vidros.
    * **RDC ANVISA nº 222/2018 (Saúde)** [🔗](https://bvsms.saude.gov.br/bvs/saudelegis/anvisa/2018/rdc0222_28_03_2018.pdf): Grupos A (Biológicos), B (Químicos), C (Radioativos), D (Comuns) e E (Perfurocortantes).
    * **Resolução CONAMA nº 307/2002 (Obras)** [🔗](http://conama.mma.gov.br/resolucoes/2002/res307.html): Classes A (Alvenaria), B (Plásticos/Madeira), C (Gesso), D (Tintas/Telhas).
    """)

    st.markdown("---")
    st.markdown("### 🌐 **Sistemas Ambientais & Órgãos Reguladores**")
    st.markdown("""
    * **MTR (Manifesto de Transporte de Resíduos)** [🔗](https://sinir.gov.br/sistemas/mtr/): Documento obrigatório para acompanhar o transporte de resíduos da origem até o destinador final. Emitido eletronicamente.
    * **SINIR** [🔗](https://sinir.gov.br/): Plataforma do Ministério do Meio Ambiente que consolida dados do MTR nacional e Inventários.
    * **SIGOR (São Paulo)** [🔗](https://cetesb.sp.gov.br/sigor/): Sistema estadual gerido pela CETESB para controle de MTR e emissão de CADRI.
    * **CETESB / Órgãos Estaduais** [🔗](https://cetesb.sp.gov.br/): Agências encarregadas do licenciamento ambiental (LP, LI, LO) e fiscalização.
    """)
